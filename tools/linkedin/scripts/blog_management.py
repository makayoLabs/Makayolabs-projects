import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create a new Excel workbook and select the active sheet
def create_excel_sheet(file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Blog Management"
    
    # Define the headers
    headers = ["Date Added", "Blog Link", "Blog Title", "Blog Summary", "LinkedIn Post", "Status"]
    
    # Set headers in the first row
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header
    
    # Set the column widths for better readability
    sheet.column_dimensions['A'].width = 15  # Date Added
    sheet.column_dimensions['B'].width = 50  # Blog Link
    sheet.column_dimensions['C'].width = 40  # Blog Title
    sheet.column_dimensions['D'].width = 100  # Blog Summary
    sheet.column_dimensions['E'].width = 80   # LinkedIn Post
    sheet.column_dimensions['F'].width = 15   # Status
    
    # Save the workbook to the specified file
    workbook.save(file_name)
    print(f"Excel sheet '{file_name}' created successfully.")

# Function to add a new blog entry
def add_blog_entry(file_name, blog_link, blog_title, blog_summary, linkedin_post, status):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    
    # Find the next empty row
    next_row = sheet.max_row + 1
    
    # Get the current date
    date_added = datetime.now().strftime("%Y-%m-%d")
    
    # Insert the data
    sheet[f"A{next_row}"] = date_added
    sheet[f"B{next_row}"] = blog_link
    sheet[f"C{next_row}"] = blog_title
    sheet[f"D{next_row}"] = blog_summary
    sheet[f"E{next_row}"] = linkedin_post
    sheet[f"F{next_row}"] = status
    
    # Save the workbook
    workbook.save(file_name)
    print(f"Blog entry added to '{file_name}'.")

# Example usage:
file_name = "blog_management_expanded.xlsx"
create_excel_sheet(file_name)

# Adding a blog entry example
blog_link = "https://example.com/my-awesome-blog"
blog_title = "My Awesome Blog"
blog_summary = "This is a summary of the awesome blog post."
linkedin_post = "Check out my latest blog on awesome topics! #blogging #awesome"
status = "Pending"

add_blog_entry(file_name, blog_link, blog_title, blog_summary, linkedin_post, status)
